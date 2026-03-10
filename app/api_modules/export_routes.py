from io import BytesIO

from flask import Response, g, request
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import select

from app.api import api_bp
from app.api_modules.common import json_error
from app.api_modules.services import get_fiscal_year_for_project, get_project_and_membership, get_project_readiness, get_transfer_rows, get_trial_balance_rows
from app.auth import require_api_session
from app.models import JournalHeader, ProjectMembership


@api_bp.get("/projects/<int:project_id>/exports/finance.xlsx")
@require_api_session
def export_project_finance_excel(project_id):
    db_session = g.db_session
    project, membership = get_project_and_membership(
        db_session,
        g.current_user.id,
        project_id,
        active_company_id=g.current_session.active_company_id,
    )
    if membership is None or project is None:
        return json_error("project_access_denied", 403)
    fiscal_year_id = request.args.get("fiscal_year_id", type=int)
    workbook = Workbook()
    summary_sheet = workbook.active
    if summary_sheet is None:
        return json_error("workbook_initialization_failed", 500)
    summary_sheet.title = "Summary"
    readiness = get_project_readiness(db_session, project.id)
    summary_sheet.append(["Project", project.code, project.name_en, project.name_ar])
    summary_sheet.append(["Currency", project.currency_code])
    summary_sheet.append(["Fiscal Years", readiness["counts"]["fiscal_years"]])
    summary_sheet.append(["Accounts", readiness["counts"]["accounts"]])
    summary_sheet.append(["Budgets", readiness["counts"]["budgets"]])
    if fiscal_year_id:
        fiscal_year = get_fiscal_year_for_project(db_session, project.id, fiscal_year_id)
        if fiscal_year is None:
            return json_error("invalid_fiscal_year")
        journals_sheet = workbook.create_sheet("Journals")
        journals_sheet.append(["Date", "Number", "Description", "Debit", "Credit"])
        journals = db_session.scalars(
            select(JournalHeader)
            .where(JournalHeader.project_id == project.id, JournalHeader.fiscal_year_id == fiscal_year.id)
            .order_by(JournalHeader.entry_date, JournalHeader.journal_number)
        ).all()
        for journal in journals:
            journals_sheet.append([journal.entry_date.isoformat(), journal.journal_number, journal.description, sum(float(line.debit) for line in journal.lines), sum(float(line.credit) for line in journal.lines)])
        tb_sheet = workbook.create_sheet("Trial Balance")
        tb_sheet.append(["Account Code", "Arabic Name", "English Name", "Debit", "Credit", "Balance"])
        rows, total_debit, total_credit = get_trial_balance_rows(db_session, project.id, fiscal_year.id)
        for row in rows:
            tb_sheet.append([row["account_code"], row["account_name_ar"], row["account_name_en"], row["debit"], row["credit"], row["balance"]])
        tb_sheet.append(["Totals", "", "", total_debit, total_credit, total_debit - total_credit])
        transfer_sheet = workbook.create_sheet("Transfers")
        transfer_sheet.append(["Date", "Direction", "Source Project", "Destination Project", "Description", "Amount"])
        for transfer in get_transfer_rows(db_session, project.id, fiscal_year.id):
            transfer_sheet.append([transfer["transfer_date"], transfer["direction"], transfer["source_project_code"], transfer["destination_project_code"], transfer["description"], transfer["total_amount"]])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={project.code.lower()}-finance.xlsx"})


@api_bp.get("/projects/<int:project_id>/exports/finance.pdf")
@require_api_session
def export_project_finance_pdf(project_id):
    db_session = g.db_session
    project, membership = get_project_and_membership(
        db_session,
        g.current_user.id,
        project_id,
        active_company_id=g.current_session.active_company_id,
    )
    if membership is None or project is None:
        return json_error("project_access_denied", 403)
    fiscal_year_id = request.args.get("fiscal_year_id", type=int)
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    y_position = 800
    readiness = get_project_readiness(db_session, project.id)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(40, y_position, f"Project Finance Export - {project.code}")
    pdf.setFont("Helvetica", 10)
    y_position -= 30
    for line in [f"Project: {project.name_en} / {project.name_ar}", f"Currency: {project.currency_code}", f"Fiscal Years: {readiness['counts']['fiscal_years']}", f"Accounts: {readiness['counts']['accounts']}", f"Budgets: {readiness['counts']['budgets']}"]:
        pdf.drawString(40, y_position, line)
        y_position -= 16
    if fiscal_year_id:
        fiscal_year = get_fiscal_year_for_project(db_session, project.id, fiscal_year_id)
        if fiscal_year is None:
            return json_error("invalid_fiscal_year")
        y_position -= 12
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(40, y_position, f"Trial Balance - {fiscal_year.code}")
        y_position -= 18
        pdf.setFont("Helvetica", 9)
        rows, total_debit, total_credit = get_trial_balance_rows(db_session, project.id, fiscal_year.id)
        for row in rows:
            pdf.drawString(40, y_position, f"{row['account_code']} | {row['account_name_en']} | D {row['debit']:.2f} | C {row['credit']:.2f}")
            y_position -= 14
            if y_position < 60:
                pdf.showPage()
                pdf.setFont("Helvetica", 9)
                y_position = 800
        pdf.drawString(40, y_position, f"Totals | D {total_debit:.2f} | C {total_credit:.2f}")
        y_position -= 24
        if y_position < 80:
            pdf.showPage()
            y_position = 800
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(40, y_position, f"Project Transfers - {fiscal_year.code}")
        y_position -= 18
        pdf.setFont("Helvetica", 9)
        for transfer in get_transfer_rows(db_session, project.id, fiscal_year.id):
            pdf.drawString(40, y_position, f"{transfer['transfer_date']} | {transfer['direction']} | {transfer['source_project_code']} -> {transfer['destination_project_code']} | {transfer['total_amount']:.2f}")
            y_position -= 14
            if y_position < 60:
                pdf.showPage()
                pdf.setFont("Helvetica", 9)
                y_position = 800
    pdf.save()
    output.seek(0)
    return Response(output.getvalue(), mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={project.code.lower()}-finance.pdf"})


@api_bp.get("/exports/projects.xlsx")
@require_api_session
def export_projects_excel():
    db_session = g.db_session
    memberships = db_session.scalars(
        select(ProjectMembership)
        .join(ProjectMembership.project)
        .where(
            ProjectMembership.user_id == g.current_user.id,
            ProjectMembership.project.has(company_id=g.current_session.active_company_id),
        )
    ).all()
    projects = [membership.project for membership in memberships]
    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        return json_error("workbook_initialization_failed", 500)
    worksheet.title = "Projects"
    worksheet.append(["ID", "Code", "Arabic Name", "English Name", "Currency", "Active"])
    for project in projects:
        worksheet.append([project.id, project.code, project.name_ar, project.name_en, project.currency_code, project.is_active])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=projects.xlsx"})


@api_bp.get("/exports/projects.pdf")
@require_api_session
def export_projects_pdf():
    db_session = g.db_session
    memberships = db_session.scalars(
        select(ProjectMembership)
        .join(ProjectMembership.project)
        .where(
            ProjectMembership.user_id == g.current_user.id,
            ProjectMembership.project.has(company_id=g.current_session.active_company_id),
        )
    ).all()
    projects = [membership.project for membership in memberships]
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    y_position = 800
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(50, y_position, "Projects Export")
    pdf.setFont("Helvetica", 10)
    y_position -= 30
    for project in projects:
        pdf.drawString(50, y_position, f"{project.code} | {project.name_en} | {project.name_ar} | {project.currency_code}")
        y_position -= 18
        if y_position < 50:
            pdf.showPage()
            y_position = 800
    pdf.save()
    output.seek(0)
    return Response(output.getvalue(), mimetype="application/pdf", headers={"Content-Disposition": "attachment; filename=projects.pdf"})